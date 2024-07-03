# excel_utils.py
import openpyxl
import time

def save_to_excel(excel_file, 
                  distance_thumb_cmc_mcp, 
                  distance_thumb_mcp_ip, 
                  distance_thumb_ip_tip, 
                  distance_wt_if_mcp, 
                  distance_if_mcp_pip, 
                  distance_if_pip_dip, 
                  distance_if_dip_tip, 
                  distance_mf_mcp_pip, 
                  distance_mf_pip_dip, 
                  distance_mf_dip_tip, 
                  distance_rf_mcp_pip, 
                  distance_rf_pip_dip, 
                  distance_rf_dip_tip, 
                  distance_pf_mcp_pip, 
                  distance_pf_pip_dip, 
                  distance_pf_dip_tip,
                  distance_wt_mf_tip,
                  distance_thumb_pf_tip
                  ):
    # Create a new workbook or open an existing one
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Write the header row
        worksheet['A1'] = 'Timestamp'
        worksheet['B1'] = 'Thumb CMC to Thumb MCP'
        worksheet['C1'] = 'Thumb MCP to Thumb IP'
        worksheet['D1'] = 'Thumb IP to Thumb Tip'
        worksheet['E1'] = 'Wrist to Index Finger MCP'
        worksheet['F1'] = 'Index Finger MCP to PIP'
        worksheet['G1'] = 'Index Finger PIP to DIP'
        worksheet['H1'] = 'Index Finger DIP to Tip'
        worksheet['I1'] = 'Middle Finger MCP to PIP'
        worksheet['J1'] = 'Middle Finger PIP to DIP'
        worksheet['K1'] = 'Middle Finger DIP to Tip'
        worksheet['L1'] = 'Ring Finger MCP to PIP'
        worksheet['M1'] = 'Ring Finger PIP to DIP'
        worksheet['N1'] = 'Ring Finger DIP to Tip'
        worksheet['O1'] = 'Pinky Finger MCP to PIP'
        worksheet['P1'] = 'Pinky Finger PIP to DIP'
        worksheet['Q1'] = 'Pinky Finger DIP to Tip'
        worksheet['R1'] = 'distance_wt_mf_tip'
        worksheet['S1'] = 'distance_thumb_pf_tip'
      
    else:
        worksheet = workbook.active

    # Write the data to the Excel file
    timestamp = int(time.time())
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=timestamp)
    worksheet.cell(row=row, column=2, value=distance_thumb_cmc_mcp)
    worksheet.cell(row=row, column=3, value=distance_thumb_mcp_ip)
    worksheet.cell(row=row, column=4, value=distance_thumb_ip_tip)
    worksheet.cell(row=row, column=5, value=distance_wt_if_mcp)
    worksheet.cell(row=row, column=6, value=distance_if_mcp_pip)
    worksheet.cell(row=row, column=7, value=distance_if_pip_dip)
    worksheet.cell(row=row, column=8, value=distance_if_dip_tip)
    worksheet.cell(row=row, column=9, value=distance_mf_mcp_pip)
    worksheet.cell(row=row, column=10, value=distance_mf_pip_dip)
    worksheet.cell(row=row, column=11, value=distance_mf_dip_tip)
    worksheet.cell(row=row, column=12, value=distance_rf_mcp_pip)
    worksheet.cell(row=row, column=13, value=distance_rf_pip_dip)
    worksheet.cell(row=row, column=14, value=distance_rf_dip_tip)
    worksheet.cell(row=row, column=15, value=distance_pf_mcp_pip)
    worksheet.cell(row=row, column=16, value=distance_pf_pip_dip)
    worksheet.cell(row=row, column=17, value=distance_pf_dip_tip)
    worksheet.cell(row=row, column=18, value=distance_wt_mf_tip)
    worksheet.cell(row=row, column=19, value=distance_thumb_pf_tip)
    #worksheet
    # Adjust column widths
    for column in range(1, 18):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 20

    # Save the Excel file
    workbook.save(excel_file)